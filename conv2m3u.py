#!/usr/bin/env python

#convert the xlsx to m3u format
##EXTINF:-1, CCTV-1
#http://10.0.0.1:4022/udp/239.49.0.1:8000


from openpyxl import load_workbook
import os

# udpxy_srv="http://10.0.0.1:4022/udp/"
udpxy_srv="rtp://"


def Conv_Xls2m3u(f_xls,debug=False):
    curpath = os.path.dirname(os.path.abspath(__file__))
    xlsx_file=os.path.join(curpath,f_xls)
    workbook = load_workbook(filename=xlsx_file,data_only=True)
    f_prefix = os.path.splitext(f_xls)[0]
    sheets = workbook.sheetnames
    sheets_cnt = len(sheets)
    print(sheets)
    output_path = os.path.join(curpath, "output")
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    for s in sheets:
        st = workbook[s]
        print(st.title)
        if sheets_cnt > 1:
            m3u_file = f_prefix+"-"+st.title+".m3u"
        else:
            m3u_file = f_prefix+".m3u"
        m3u_filepath = os.path.join(output_path,m3u_file)
        h_m3u= open(m3u_filepath,"wt")
        start_idx=2
        # A2 B2 C2
        while True:
            idxA = "A"+str(start_idx)
            idxB = "B"+str(start_idx)
            idxC = "C"+str(start_idx)
            title = st[idxA].value
            if title!=None:
                if title == "#" and not debug:
                    start_idx+=1
                    continue
                igmp = str(st[idxB].value)
                if igmp[0].isdecimal():
                    print(igmp[0])
                    igmp = str(st[idxC].value)
                print(title,igmp)
                h_m3u.writelines("#EXTINF:-1, "+str(title)+"\n")
                udpxy_addr=igmp.replace("igmp://",udpxy_srv)
                h_m3u.writelines(udpxy_addr+"\n")
                start_idx+=1
            else:
                break
        h_m3u.close()

Conv_Xls2m3u("江苏电信-组播.xlsx")
Conv_Xls2m3u("江苏电信-组播(调试).xlsx",debug=True)